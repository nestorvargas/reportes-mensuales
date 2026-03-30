"""
Generates Excel (.xlsx) and PDF reports from time entry data.
"""
import io
from collections import defaultdict
from datetime import datetime

# ── openpyxl ──────────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── reportlab ─────────────────────────────────────────────────────────────
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable
)

# ── palette ───────────────────────────────────────────────────────────────
NAVY   = "0F3460"
RED    = "E94560"
LIGHT  = "EBF3FB"
GRAY   = "64748B"
WHITE  = "FFFFFF"
BORDER_COLOR = "CBD5E1"


def _hhmm(minutes: int) -> str:
    h, m = divmod(minutes, 60)
    return f"{h}:{m:02d}"


def _fmt_date(d: str) -> str:
    """'2026-03-02' → '02/Mar'"""
    try:
        dt = datetime.strptime(d, "%Y-%m-%d")
        return dt.strftime("%d/%b")
    except Exception:
        return d


# ══════════════════════════════════════════════════════════════════════════
# EXCEL
# ══════════════════════════════════════════════════════════════════════════

def _xl_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _xl_border(color: str = BORDER_COLOR) -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _xl_header_style(ws, row, col, value, bg=NAVY, fg=WHITE, bold=True, wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = _xl_fill(bg)
    cell.font = Font(color=fg, bold=bold, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    cell.border = _xl_border("FFFFFF" if bg == NAVY else BORDER_COLOR)
    return cell


def _xl_data_style(ws, row, col, value, bg=WHITE, bold=False, align="left", num_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = _xl_fill(bg)
    cell.font = Font(bold=bold, size=9)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = _xl_border()
    if num_format:
        cell.number_format = num_format
    return cell


def build_excel(date_from: str, date_to: str, rows: list[dict], summary: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    _xl_sheet_resumen(wb, date_from, date_to, summary, rows)
    _xl_sheet_pivot(wb, rows)
    _xl_sheet_detalle(wb, rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xl_sheet_resumen(wb, date_from, date_to, summary, rows):
    ws = wb.create_sheet("Resumen")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "COMPLEMENTO 360 — Reporte de Horas"
    t.fill = _xl_fill(NAVY)
    t.font = Font(color=WHITE, bold=True, size=14)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    p = ws["A2"]
    p.value = f"Período: {date_from}  →  {date_to}"
    p.fill = _xl_fill(RED)
    p.font = Font(color=WHITE, bold=True, size=10)
    p.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # KPI row
    ws.row_dimensions[3].height = 6
    kpis = [
        ("Total Horas", _hhmm(summary["total_minutes"])),
        ("Registros", str(len(rows))),
        ("Semanas", str(len(summary["by_week"]))),
        ("Work Items", str(len(summary["by_work_item"]))),
    ]
    for i, (label, val) in enumerate(kpis):
        col = i * 2 + 1
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
        _xl_header_style(ws, 4, col, label, bg=LIGHT, fg=NAVY, bold=True)
        ws.row_dimensions[4].height = 14
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
        c = ws.cell(row=5, column=col, value=val)
        c.fill = _xl_fill(WHITE)
        c.font = Font(bold=True, size=16, color=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _xl_border()
        ws.row_dimensions[5].height = 28

    ws.row_dimensions[6].height = 10

    # By type table
    row = 7
    _xl_header_style(ws, row, 1, "Tipo de Actividad", bg=NAVY)
    _xl_header_style(ws, row, 2, "Tiempo", bg=NAVY)
    _xl_header_style(ws, row, 3, "Minutos", bg=NAVY)
    _xl_header_style(ws, row, 4, "% del Total", bg=NAVY)
    ws.row_dimensions[row].height = 16

    total = summary["total_minutes"] or 1
    for i, t in enumerate(summary["by_type"]):
        row += 1
        bg = LIGHT if i % 2 == 0 else WHITE
        _xl_data_style(ws, row, 1, t["type"], bg=bg)
        _xl_data_style(ws, row, 2, _hhmm(t["minutes"]), bg=bg, align="center", bold=True)
        _xl_data_style(ws, row, 3, t["minutes"], bg=bg, align="right")
        pct = round(t["minutes"] / total * 100, 1)
        _xl_data_style(ws, row, 4, f"{pct}%", bg=bg, align="center")
        ws.row_dimensions[row].height = 15

    row += 2

    # By week table
    _xl_header_style(ws, row, 1, "Semana", bg=NAVY)
    _xl_header_style(ws, row, 2, "Tiempo", bg=NAVY)
    _xl_header_style(ws, row, 3, "Minutos", bg=NAVY)
    ws.row_dimensions[row].height = 16

    for i, w in enumerate(summary["by_week"]):
        row += 1
        bg = LIGHT if i % 2 == 0 else WHITE
        _xl_data_style(ws, row, 1, w["week"], bg=bg)
        _xl_data_style(ws, row, 2, _hhmm(w["minutes"]), bg=bg, align="center", bold=True)
        _xl_data_style(ws, row, 3, w["minutes"], bg=bg, align="right")
        ws.row_dimensions[row].height = 15

    # Column widths
    for col, w in [(1, 40), (2, 12), (3, 12), (4, 12), (5, 12), (6, 12)]:
        ws.column_dimensions[get_column_letter(col)].width = w


def _xl_sheet_pivot(wb, rows):
    """Pivot: Work Item × Day (like the screenshot)."""
    ws = wb.create_sheet("Por Día")
    ws.sheet_view.showGridLines = False

    # Collect unique dates and work items
    dates = sorted({r["date"] for r in rows})
    items = {}  # title -> {date -> minutes, week -> str}
    for r in rows:
        t = r["work_item_title"]
        if t not in items:
            items[t] = {"type": r["type"], "dates": defaultdict(int), "weeks": set()}
        items[t]["dates"][r["date"]] += r["minutes"]
        items[t]["weeks"].add(r["week"])

    # Group dates by week for column headers
    week_dates: dict[str, list[str]] = defaultdict(list)
    date_week: dict[str, str] = {}
    for r in rows:
        date_week[r["date"]] = r["week"]
    for d in dates:
        week_dates[date_week[d]].append(d)
    weeks = sorted(week_dates.keys())

    # Build header rows
    # Row 1: title
    total_cols = 2 + len(dates) + len(weeks) + 1  # type + item + dates + week totals + grand total
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    t = ws["A1"]
    t.value = "Registro de Horas por Día — Complemento 360"
    t.fill = _xl_fill(NAVY)
    t.font = Font(color=WHITE, bold=True, size=12)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Row 2: "Day of week" spanning all date columns
    col_offset = 3  # col 1=Type, col2=Work Item, col3+ = dates
    ws.merge_cells(start_row=2, start_column=col_offset, end_row=2, end_column=col_offset + len(dates) - 1)
    c = ws.cell(row=2, column=col_offset, value="Day of week")
    c.fill = _xl_fill(LIGHT)
    c.font = Font(bold=True, size=9, color=NAVY)
    c.alignment = Alignment(horizontal="center")

    # Row 3: fixed headers + date headers grouped by week
    _xl_header_style(ws, 3, 1, "Type", bg=NAVY)
    _xl_header_style(ws, 3, 2, "Work Item", bg=NAVY)
    ws.row_dimensions[3].height = 20

    date_col_map = {}
    col = col_offset
    week_start_cols = {}
    for week in weeks:
        week_start_cols[week] = col
        for d in sorted(week_dates[week]):
            dt = datetime.strptime(d, "%Y-%m-%d")
            label = dt.strftime("%d %a")  # "02 Mon"
            _xl_header_style(ws, 3, col, label, bg=NAVY)
            date_col_map[d] = col
            col += 1
        # Week total column
        _xl_header_style(ws, 3, col, week, bg=RED)
        date_col_map[f"__week_{week}"] = col
        col += 1

    # Grand total
    _xl_header_style(ws, 3, col, "Total", bg=RED)
    grand_total_col = col

    # Data rows
    sorted_items = sorted(items.items(), key=lambda x: -sum(x[1]["dates"].values()))
    for i, (title, data) in enumerate(sorted_items):
        row = 4 + i
        bg = LIGHT if i % 2 == 0 else WHITE
        _xl_data_style(ws, row, 1, data["type"], bg=bg)
        _xl_data_style(ws, row, 2, title, bg=bg)

        row_total = 0
        for week in weeks:
            week_total = 0
            for d in sorted(week_dates[week]):
                mins = data["dates"].get(d, 0)
                c = _xl_data_style(ws, row, date_col_map[d], _hhmm(mins) if mins else "", bg=bg, align="center")
                week_total += mins
                row_total += mins
            wc = _xl_data_style(ws, row, date_col_map[f"__week_{week}"],
                                 _hhmm(week_total) if week_total else "", bg=RED if week_total else bg,
                                 bold=True, align="center")
            if week_total:
                wc.font = Font(bold=True, color=WHITE, size=9)
                wc.fill = _xl_fill(RED)

        tc = _xl_data_style(ws, row, grand_total_col, _hhmm(row_total) if row_total else "",
                             bg=NAVY, bold=True, align="center")
        tc.font = Font(bold=True, color=WHITE, size=9)
        ws.row_dimensions[row].height = 18

    # Totals row
    total_row = 4 + len(sorted_items)
    ws.row_dimensions[total_row].height = 18
    _xl_data_style(ws, total_row, 1, "", bg=NAVY)
    c = ws.cell(row=total_row, column=2, value="Totals")
    c.fill = _xl_fill(NAVY)
    c.font = Font(bold=True, color=WHITE, size=10)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = _xl_border()

    for d in dates:
        col = date_col_map[d]
        total_min = sum(data["dates"].get(d, 0) for _, data in sorted_items)
        tc = _xl_data_style(ws, total_row, col, _hhmm(total_min) if total_min else "", bg=NAVY, bold=True, align="center")
        tc.font = Font(bold=True, color=WHITE, size=9)

    for week in weeks:
        wt = sum(data["dates"].get(d, 0) for _, data in sorted_items for d in sorted(week_dates[week]))
        tc = _xl_data_style(ws, total_row, date_col_map[f"__week_{week}"], _hhmm(wt) if wt else "", bg=RED, bold=True, align="center")
        tc.font = Font(bold=True, color=WHITE, size=9)

    grand = sum(sum(data["dates"].values()) for _, data in sorted_items)
    tc = _xl_data_style(ws, total_row, grand_total_col, _hhmm(grand), bg=RED, bold=True, align="center")
    tc.font = Font(bold=True, color=WHITE, size=9)

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 45
    for c_idx in range(col_offset, grand_total_col + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 9


def _xl_sheet_detalle(wb, rows):
    ws = wb.create_sheet("Detalle")
    ws.sheet_view.showGridLines = False

    headers = ["Fecha", "Semana", "Work Item", "Tipo", "Comentario", "Tiempo"]
    for ci, h in enumerate(headers, 1):
        _xl_header_style(ws, 1, ci, h)
    ws.row_dimensions[1].height = 18

    for i, r in enumerate(rows):
        row = i + 2
        bg = LIGHT if i % 2 == 0 else WHITE
        _xl_data_style(ws, row, 1, r["date"], bg=bg, align="center")
        _xl_data_style(ws, row, 2, r["week"], bg=bg, align="center")
        _xl_data_style(ws, row, 3, r["work_item_title"], bg=bg)
        _xl_data_style(ws, row, 4, r["type"], bg=bg)
        _xl_data_style(ws, row, 5, r["comment"], bg=bg)
        _xl_data_style(ws, row, 6, _hhmm(r["minutes"]), bg=bg, align="center", bold=True)
        ws.row_dimensions[row].height = 15

    widths = [12, 12, 50, 25, 55, 10]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ══════════════════════════════════════════════════════════════════════════
# PDF
# ══════════════════════════════════════════════════════════════════════════

def _rl_color(hex_str: str):
    h = hex_str.lstrip("#")
    return colors.HexColor(f"#{h}")


def build_pdf(date_from: str, date_to: str, rows: list[dict], summary: dict) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    navy = _rl_color(NAVY)
    red  = _rl_color(RED)

    title_style = ParagraphStyle("title", fontSize=16, textColor=colors.white,
                                  fontName="Helvetica-Bold", alignment=1)
    sub_style   = ParagraphStyle("sub",   fontSize=10, textColor=colors.white,
                                  fontName="Helvetica", alignment=1)
    h2_style    = ParagraphStyle("h2",    fontSize=11, textColor=navy,
                                  fontName="Helvetica-Bold", spaceAfter=4)
    cell_style  = ParagraphStyle("cell",  fontSize=8,  fontName="Helvetica", leading=10)

    elements = []

    # ── Header block ──
    header_data = [[
        Paragraph("COMPLEMENTO 360 — Reporte de Horas", title_style),
    ], [
        Paragraph(f"Período: {date_from}  →  {date_to}", sub_style),
    ]]
    header_tbl = Table(header_data, colWidths=["100%"])
    header_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), navy),
        ("BACKGROUND", (0, 1), (-1, 1), red),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [navy, red]),
    ]))
    elements.append(header_tbl)
    elements.append(Spacer(1, 0.4 * cm))

    # ── KPI cards ──
    total = summary["total_minutes"]
    kpi_data = [[
        _kpi_cell("Total Horas", _hhmm(total)),
        _kpi_cell("Registros", str(len(rows))),
        _kpi_cell("Semanas", str(len(summary["by_week"]))),
        _kpi_cell("Work Items", str(len(summary["by_work_item"]))),
        _kpi_cell("Días trabajados", str(len({r["date"] for r in rows}))),
    ]]
    kpi_tbl = Table(kpi_data, colWidths=[None] * 5)
    kpi_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), _rl_color(LIGHT)),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.white),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(kpi_tbl)
    elements.append(Spacer(1, 0.5 * cm))

    # ── Two-column: by type + by week ──
    type_block  = _pdf_summary_table("Horas por Tipo",   ["Tipo", "Tiempo", "%"],
                                      [[t["type"], _hhmm(t["minutes"]),
                                        f"{round(t['minutes']/total*100,1)}%"]
                                       for t in summary["by_type"]])
    week_block  = _pdf_summary_table("Horas por Semana", ["Semana", "Tiempo"],
                                      [[w["week"], _hhmm(w["minutes"])]
                                       for w in summary["by_week"]])

    two_col = Table([[type_block, week_block]], colWidths=["60%", "40%"])
    two_col.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(two_col)
    elements.append(Spacer(1, 0.5 * cm))

    # ── Work items table ──
    elements.append(Paragraph("Horas por Work Item", h2_style))
    wi_data = [["Work Item", "Tiempo", "%"]] + [
        [Paragraph(w["title"], cell_style), _hhmm(w["minutes"]),
         f"{round(w['minutes']/total*100,1)}%"]
        for w in summary["by_work_item"]
    ]
    wi_tbl = Table(wi_data, colWidths=["75%", "12%", "13%"])
    wi_tbl.setStyle(_base_table_style(navy, red))
    elements.append(wi_tbl)
    elements.append(Spacer(1, 0.5 * cm))

    # ── Detail table ──
    elements.append(Paragraph("Detalle de Registros", h2_style))
    detail_data = [["Fecha", "Semana", "Work Item", "Tipo", "Comentario", "Tiempo"]] + [
        [r["date"], r["week"],
         Paragraph(r["work_item_title"], cell_style),
         Paragraph(r["type"], cell_style),
         Paragraph(r["comment"], cell_style),
         _hhmm(r["minutes"])]
        for r in rows
    ]
    det_tbl = Table(detail_data, colWidths=["9%", "9%", "25%", "15%", "33%", "9%"])
    det_tbl.setStyle(_base_table_style(navy, red))
    elements.append(det_tbl)

    doc.build(elements)
    return buf.getvalue()


def _kpi_cell(label: str, value: str):
    return Paragraph(
        f'<para align="center"><font size="8" color="#{GRAY}">{label}</font><br/>'
        f'<font size="18" color="#{NAVY}"><b>{value}</b></font></para>',
        ParagraphStyle("kpi", leading=22)
    )


def _pdf_summary_table(title: str, headers: list, data_rows: list):
    from reportlab.platypus import KeepInFrame
    styles = getSampleStyleSheet()
    h2 = ParagraphStyle("h2b", fontSize=11, textColor=_rl_color(NAVY),
                         fontName="Helvetica-Bold", spaceAfter=4)
    tdata = [headers] + data_rows
    tbl = Table(tdata)
    tbl.setStyle(_base_table_style(_rl_color(NAVY), _rl_color(RED)))
    return [Paragraph(title, h2), tbl]


def _base_table_style(header_bg, alt_bg):
    light = _rl_color(LIGHT)
    return TableStyle([
        # Header
        ("BACKGROUND",   (0, 0), (-1, 0),  header_bg),
        ("TEXTCOLOR",    (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",     (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0), (-1, 0),  8),
        ("ALIGN",        (0, 0), (-1, 0),  "CENTER"),
        ("TOPPADDING",   (0, 0), (-1, 0),  5),
        ("BOTTOMPADDING",(0, 0), (-1, 0),  5),
        # Data
        ("FONTSIZE",     (0, 1), (-1, -1), 7.5),
        ("TOPPADDING",   (0, 1), (-1, -1), 3),
        ("BOTTOMPADDING",(0, 1), (-1, -1), 3),
        ("ROWBACKGROUNDS",(0, 1),(-1, -1), [light, colors.white]),
        ("GRID",         (0, 0), (-1, -1), 0.3, _rl_color(BORDER_COLOR)),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
    ])
